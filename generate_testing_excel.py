import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# ============================================================
# STYLES
# ============================================================
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='1B4F72', end_color='1B4F72', fill_type='solid')
sub_header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
sub_header_fill = PatternFill(start_color='2E86C1', end_color='2E86C1', fill_type='solid')
section_font = Font(name='Calibri', bold=True, size=11, color='1B4F72')
section_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
day_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
day_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
normal_font = Font(name='Calibri', size=10)
bold_font = Font(name='Calibri', bold=True, size=10)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

guru_fill = PatternFill(start_color='E8F8F5', end_color='E8F8F5', fill_type='solid')
wali_fill = PatternFill(start_color='F5EEF8', end_color='F5EEF8', fill_type='solid')
admin_fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
crosscheck_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
edge_fill = PatternFill(start_color='FEF9E7', end_color='FEF9E7', fill_type='solid')
recap_fill = PatternFill(start_color='FDEBD0', end_color='FDEBD0', fill_type='solid')

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

def style_day_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = day_font
        cell.fill = day_fill
        cell.alignment = center
        cell.border = thin_border

def style_section(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = left_wrap
        cell.border = thin_border

def apply_row(ws, row, max_col, font=None, fill=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        if font: cell.font = font
        if fill: cell.fill = fill
        cell.alignment = left_wrap


# ============================================================
# SHEET 1: KONFIGURASI TESTER
# ============================================================
ws1 = wb.active
ws1.title = "Konfigurasi Tester"
ws1.sheet_properties.tabColor = "1B4F72"

ws1.merge_cells('A1:G1')
ws1.cell(row=1, column=1, value="KONFIGURASI TESTER - SKENARIO TESTING 14 HARI").font = Font(name='Calibri', bold=True, size=14, color='1B4F72')
ws1.cell(row=1, column=1).alignment = center
ws1.row_dimensions[1].height = 30

ws1.merge_cells('A2:G2')
ws1.cell(row=2, column=1, value="12 Tester | T1-T9: Guru+Wali | T10-T12: Admin+Guru+Wali | 2 Sekolah | ~30 menit/hari").font = Font(name='Calibri', italic=True, size=11, color='5D6D7E')
ws1.cell(row=2, column=1).alignment = center

# ---- GROUP A: T1-T9 (Guru + Wali) ----
ws1.merge_cells('A4:G4')
ws1.cell(row=4, column=1, value="GRUP A: GURU + WALI (T1 - T9)").font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
ws1.cell(row=4, column=1).fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
ws1.cell(row=4, column=1).alignment = center
for c in range(1, 8):
    ws1.cell(row=4, column=c).border = thin_border

headers = ['No', 'Tester', 'Role di Sekolah A', 'Role di Sekolah B', 'Mata Pelajaran', 'Keterangan', 'Fokus Khusus']
for col, h in enumerate(headers, 1):
    ws1.cell(row=5, column=col, value=h)
style_header_row(ws1, 5, 7)

testers_gw = [
    ['T1', 'Tester 1', 'GURU (Wali Kelas)', 'WALI (1 anak)', 'Matematika', 'Guru homeroom Sekolah A', 'AI Rekomendasi'],
    ['T2', 'Tester 2', 'GURU', 'WALI (1 anak)', 'Bahasa Indonesia', 'Guru mapel bahasa', 'CRUD Aktivitas'],
    ['T3', 'Tester 3', 'GURU', 'WALI (2 anak)', 'IPA', 'Guru mapel sains', 'Multi-anak (Wali B)'],
    ['T4', 'Tester 4', 'GURU', 'WALI (1 anak)', 'IPS', 'Guru mapel sosial', 'Input Nilai bulk'],
    ['T5', 'Tester 5', 'WALI (1 anak)', 'GURU (Wali Kelas)', 'Matematika', 'Guru homeroom Sekolah B', 'AI Rekomendasi'],
    ['T6', 'Tester 6', 'WALI (2 anak)', 'GURU', 'B. Inggris', 'Multi-anak (Wali A)', 'Multi-anak switching'],
    ['T7', 'Tester 7', 'WALI (1 anak)', 'GURU', 'PKN', 'Guru mapel umum', 'Edge cases'],
    ['T8', 'Tester 8', 'GURU', 'WALI (1 anak)', 'Seni Budaya', 'Guru mapel seni', 'Materi & RPP AI'],
    ['T9', 'Tester 9', 'WALI (1 anak)', 'GURU', 'Penjaskes', 'Guru mapel olahraga', 'Absensi & Raport'],
]

gw_fills = [guru_fill]*4 + [wali_fill]*3 + [guru_fill, wali_fill]
for i, t in enumerate(testers_gw):
    row = 6 + i
    for col, val in enumerate(t, 1):
        ws1.cell(row=row, column=col, value=val)
    apply_row(ws1, row, 7, font=normal_font, fill=gw_fills[i])
    ws1.cell(row=row, column=1).alignment = center

# ---- GROUP B: T10-T12 (Admin + Guru + Wali) ----
row_admin_start = 17
ws1.merge_cells(f'A{row_admin_start}:G{row_admin_start}')
ws1.cell(row=row_admin_start, column=1, value="GRUP B: ADMIN + GURU + WALI (T10 - T12)").font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
ws1.cell(row=row_admin_start, column=1).fill = PatternFill(start_color='2E86C1', end_color='2E86C1', fill_type='solid')
ws1.cell(row=row_admin_start, column=1).alignment = center
for c in range(1, 8):
    ws1.cell(row=row_admin_start, column=c).border = thin_border

headers_admin = ['No', 'Tester', 'Role di Sekolah A', 'Role di Sekolah B', 'Fokus Admin', 'Fokus Guru', 'Fokus Wali']
r = row_admin_start + 1
for col, h in enumerate(headers_admin, 1):
    ws1.cell(row=r, column=col, value=h)
style_header_row(ws1, r, 7)

testers_admin = [
    ['T10', 'Tester 10', 'ADMIN + GURU (B. Arab)', 'WALI (1 anak)', 'Data Master (Siswa, Guru, Kelas)', 'CRUD Materi & Nilai', 'Monitoring anak'],
    ['T11', 'Tester 11', 'WALI (1 anak)', 'ADMIN + GURU (TIK)', 'Jadwal, Pengumuman, Absensi', 'RPP & Raport', 'Monitoring anak'],
    ['T12', 'Tester 12', 'ADMIN', 'ADMIN', 'Keuangan, Settings, RPP Approval', 'N/A (Admin only)', 'N/A (Admin only)'],
]

for i, t in enumerate(testers_admin):
    r = row_admin_start + 2 + i
    for col, val in enumerate(t, 1):
        ws1.cell(row=r, column=col, value=val)
    apply_row(ws1, r, 7, font=normal_font, fill=admin_fill)
    ws1.cell(row=r, column=1).alignment = center

# Ringkasan 14 hari
row = 24
ws1.merge_cells(f'A{row}:G{row}')
ws1.cell(row=row, column=1, value="RINGKASAN JADWAL 14 HARI").font = Font(name='Calibri', bold=True, size=12, color='1B4F72')

schedule_headers = ['Hari', 'Tema', 'T1-T9 (Guru+Wali)', 'T10-T11 (Admin+Guru+Wali)', 'T12 (Admin)', 'Estimasi']
row += 1
for col, h in enumerate(schedule_headers, 1):
    ws1.cell(row=row, column=col, value=h)
style_header_row(ws1, row, 6)

schedule_summary = [
    ['Hari 1', 'Login, Dashboard & Pengumuman', 'Guru+Wali Dashboard', 'Admin+Guru+Wali Dashboard', 'Admin Dashboard', '~30 mnt'],
    ['Hari 2', 'Jadwal & Navigasi / Data Siswa', 'Guru: Jadwal Mengajar', 'Admin: Manajemen Siswa', 'Admin: Manajemen Siswa', '~30 mnt'],
    ['Hari 3', 'Aktivitas Kelas / Data Guru', 'Guru: Buat Aktivitas', 'Admin: Manajemen Guru', 'Admin: Manajemen Guru', '~30 mnt'],
    ['Hari 4', 'Kelola Aktivitas / Data Kelas', 'Guru: Edit/Hapus/Filter', 'Admin: Manajemen Kelas', 'Admin: Manajemen Kelas', '~30 mnt'],
    ['Hari 5', 'Absensi / Data Mapel & Jadwal', 'Guru: Input Absensi', 'Admin: Mapel & Jadwal', 'Admin: Mapel & Jadwal', '~30 mnt'],
    ['Hari 6', 'Materi CRUD / Pengumuman Admin', 'Guru: CRUD Materi', 'Admin: Kelola Pengumuman', 'Admin: Kelola Pengumuman', '~30 mnt'],
    ['Hari 7', 'Materi AI & RPP / Absensi Report', 'Guru: AI Materi + RPP', 'Admin: Laporan Absensi', 'Admin: Laporan Absensi', '~30 mnt'],
    ['Hari 8', 'RPP AI / RPP Approval & Nilai', 'Guru: RPP AI + Export', 'Admin: RPP Approval + Nilai', 'Admin: RPP Approval', '~30 mnt'],
    ['Hari 9', 'Input Nilai / Raport & Keuangan', 'Guru: Input Nilai + Rekap', 'Admin: Raport Publish', 'Admin: Keuangan (Jenis)', '~30 mnt'],
    ['Hari 10', 'Raport & Rekomendasi / Keuangan', 'Guru: Raport + AI Rekom', 'Admin: Keuangan (Tagihan)', 'Admin: Keuangan (Tagihan)', '~30 mnt'],
    ['Hari 11', 'Wali: Aktivitas & Nilai / Keuangan', 'Wali: Aktivitas + Nilai', 'Admin: Keuangan (Bayar) + Wali', 'Admin: Keuangan (Bayar)', '~30 mnt'],
    ['Hari 12', 'Wali: Kehadiran & Billing / Settings', 'Wali: Kehadiran + Billing', 'Admin: Settings Sekolah + Wali', 'Admin: Settings Sekolah', '~30 mnt'],
    ['Hari 13', 'Cross-check Semua Role', 'Wali: E-Raport + Cross-check', 'Admin+Guru+Wali Cross-check', 'Admin Cross-check', '~30 mnt'],
    ['Hari 14', 'Edge Cases & Rekapitulasi', 'Edge Cases + Rekap', 'Edge Cases + Rekap', 'Edge Cases + Rekap', '~30 mnt'],
]

for d in schedule_summary:
    row += 1
    for col, val in enumerate(d, 1):
        ws1.cell(row=row, column=col, value=val)
    apply_row(ws1, row, 6, font=normal_font)
    ws1.cell(row=row, column=1).font = bold_font
    ws1.cell(row=row, column=1).alignment = center

ws1.column_dimensions['A'].width = 10
ws1.column_dimensions['B'].width = 14
ws1.column_dimensions['C'].width = 36
ws1.column_dimensions['D'].width = 36
ws1.column_dimensions['E'].width = 36
ws1.column_dimensions['F'].width = 30
ws1.column_dimensions['G'].width = 24


# ============================================================
# SHEET 2: JADWAL 14 HARI — GURU + WALI (T1-T9)
# ============================================================
ws2 = wb.create_sheet("Jadwal Guru+Wali (T1-T9)")
ws2.sheet_properties.tabColor = "27AE60"

cols_gw = ['No', 'Role', 'Kategori', 'Task Testing',
           'T1', 'T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'T8', 'T9', 'Catatan Bug']
MAX_GW = len(cols_gw)
for col, h in enumerate(cols_gw, 1):
    ws2.cell(row=1, column=col, value=h)
style_header_row(ws2, 1, MAX_GW)

days_gw = [
    ("HARI 1 — Login, Dashboard & Pengumuman (~30 mnt)", [
        ("GURU (Sekolah masing-masing)", guru_fill, [
            ('Guru', 'Login', 'Login dengan akun Guru di sekolah masing-masing'),
            ('Guru', 'Dashboard', 'Verifikasi dashboard Guru tampil dengan benar'),
            ('Guru', 'Dashboard', 'Cek statistik: total siswa, kelas mengajar, jadwal hari ini'),
            ('Guru', 'Dashboard', 'Cek badge unread pada setiap menu'),
            ('Guru', 'Pengumuman', 'Buka & baca pengumuman terbaru'),
            ('Guru', 'Pengumuman', 'Verifikasi badge unread hilang setelah dibaca'),
            ('Guru', 'Settings', 'Ganti bahasa (ID ↔ EN), verifikasi teks berubah'),
            ('Guru', 'Tour', 'Cek interactive tour/tutorial muncul'),
        ]),
        ("WALI (Sekolah masing-masing)", wali_fill, [
            ('Wali', 'Login', 'Login/switch ke akun Wali di sekolah lainnya'),
            ('Wali', 'Dashboard', 'Verifikasi dashboard Wali tampil dengan benar'),
            ('Wali', 'Dashboard', 'Cek jumlah anak terdaftar'),
            ('Wali', 'Dashboard', 'Cek badge unread tiap menu'),
            ('Wali', 'Multi-Anak', 'T3 & T6: Switch antar anak, verifikasi data berubah'),
            ('Wali', 'Pengumuman', 'Buka & baca pengumuman'),
            ('Wali', 'Pengumuman', 'Verifikasi badge unread hilang setelah dibaca'),
            ('Wali', 'Settings', 'Ganti bahasa (ID ↔ EN)'),
        ]),
    ]),
    ("HARI 2 — Jadwal Mengajar & Navigasi (~30 mnt)", [
        ("GURU — Jadwal Mengajar", guru_fill, [
            ('Guru', 'Jadwal', 'Buka halaman Jadwal Mengajar'),
            ('Guru', 'Jadwal', 'Verifikasi jadwal hari ini tampil benar'),
            ('Guru', 'Jadwal', 'Cek jadwal minggu ini'),
            ('Guru', 'Jadwal', 'Verifikasi data jadwal sesuai mapel yang diajar'),
            ('Guru', 'Navigasi', 'Buka setiap menu, verifikasi navigasi lancar'),
            ('Guru', 'Navigasi', 'Klik back/forward di setiap halaman'),
            ('Guru', 'Navigasi', 'Buka notifikasi, verifikasi list tampil'),
            ('Guru', 'Dashboard', 'Cek overview materi, RPP, attendance di dashboard'),
        ]),
    ]),
    ("HARI 3 — Aktivitas Kelas: Buat & Tipe (~30 mnt)", [
        ("GURU — Buat Aktivitas Kelas", guru_fill, [
            ('Guru', 'Aktivitas', 'Buka halaman Aktivitas Kelas'),
            ('Guru', 'Aktivitas', 'Buat aktivitas kelas tipe UMUM (judul, deskripsi, kelas)'),
            ('Guru', 'Aktivitas', 'Buat aktivitas tipe SISWA TERTENTU — pilih beberapa siswa'),
            ('Guru', 'Aktivitas', 'Buat aktivitas tipe TUGAS/PR — set deadline'),
            ('Guru', 'Aktivitas', 'Buat aktivitas dengan LINK MEETING ONLINE'),
            ('Guru', 'Aktivitas', 'Verifikasi semua aktivitas muncul di list'),
            ('Guru', 'Aktivitas', 'Buka detail setiap aktivitas, verifikasi data benar'),
            ('Guru', 'Aktivitas', 'Verifikasi tipe aktivitas tampil dengan ikon/label benar'),
        ]),
    ]),
    ("HARI 4 — Aktivitas Kelas: Edit, Hapus, Filter (~30 mnt)", [
        ("GURU — Kelola Aktivitas Kelas", guru_fill, [
            ('Guru', 'Aktivitas', 'Edit salah satu aktivitas — ubah judul & deskripsi'),
            ('Guru', 'Aktivitas', 'Edit aktivitas — ubah tipe (misal umum → tugas)'),
            ('Guru', 'Aktivitas', 'Hapus salah satu aktivitas, verifikasi hilang dari list'),
            ('Guru', 'Filter', 'Filter aktivitas berdasarkan TANGGAL'),
            ('Guru', 'Filter', 'Filter aktivitas berdasarkan KELAS'),
            ('Guru', 'Filter', 'Filter aktivitas berdasarkan MATA PELAJARAN'),
            ('Guru', 'Filter', 'Kombinasi filter (tanggal + kelas)'),
            ('Guru', 'Filter', 'Reset filter, verifikasi semua data muncul kembali'),
        ]),
    ]),
    ("HARI 5 — Absensi Siswa (~30 mnt)", [
        ("GURU — Input & Kelola Absensi", guru_fill, [
            ('Guru', 'Absensi', 'Buka halaman Absensi/Presensi'),
            ('Guru', 'Absensi', 'Pilih kelas yang diajar'),
            ('Guru', 'Absensi', 'Input: tandai beberapa siswa HADIR'),
            ('Guru', 'Absensi', 'Input: tandai 1-2 siswa SAKIT'),
            ('Guru', 'Absensi', 'Input: tandai 1 siswa IZIN'),
            ('Guru', 'Absensi', 'Input: tandai 1 siswa ALPHA'),
            ('Guru', 'Absensi', 'Simpan data absensi'),
            ('Guru', 'Absensi', 'Verifikasi rekap: jumlah H/S/I/A benar'),
            ('Guru', 'Absensi', 'Buka detail absensi per siswa'),
            ('Guru', 'Absensi', 'Filter absensi berdasarkan rentang tanggal'),
            ('Guru', 'Absensi', 'Ubah status absensi siswa (misal Hadir → Sakit)'),
            ('Guru', 'Absensi', 'Simpan ulang & verifikasi perubahan tersimpan'),
        ]),
    ]),
    ("HARI 6 — Materi Pembelajaran CRUD (~30 mnt)", [
        ("GURU — Kelola Materi", guru_fill, [
            ('Guru', 'Materi', 'Buka halaman Materi Pembelajaran'),
            ('Guru', 'Materi', 'Buat materi baru — isi judul, bab, sub-bab, konten'),
            ('Guru', 'Materi', 'Verifikasi materi baru muncul di list'),
            ('Guru', 'Materi', 'Buka detail materi, verifikasi konten lengkap'),
            ('Guru', 'Materi', 'Edit materi — ubah judul dan konten'),
            ('Guru', 'Materi', 'Verifikasi perubahan tersimpan'),
            ('Guru', 'Materi', 'Hapus materi, verifikasi hilang dari list'),
            ('Guru', 'Materi', 'Buat 2-3 materi tambahan untuk test AI besok'),
        ]),
    ]),
    ("HARI 7 — Materi AI & RPP Manual (~30 mnt)", [
        ("GURU — AI Materi & RPP", guru_fill, [
            ('Guru', 'Materi AI', 'Buka fitur Generate Materi dengan AI'),
            ('Guru', 'Materi AI', 'Masukkan prompt untuk generate materi'),
            ('Guru', 'Materi AI', 'Tunggu proses, verifikasi loading indicator'),
            ('Guru', 'Materi AI', 'Verifikasi hasil AI materi tampil lengkap'),
            ('Guru', 'Materi AI', 'Regenerate materi AI dengan prompt berbeda'),
            ('Guru', 'Materi AI', 'Verifikasi hasil baru berbeda dari sebelumnya'),
            ('Guru', 'RPP', 'Buka halaman RPP'),
            ('Guru', 'RPP', 'Buat RPP baru secara MANUAL (isi semua field)'),
            ('Guru', 'RPP', 'Verifikasi RPP tersimpan di list'),
            ('Guru', 'RPP', 'Lihat detail RPP yang baru dibuat'),
            ('Guru', 'RPP', 'Edit RPP — ubah beberapa field'),
        ]),
    ]),
    ("HARI 8 — RPP AI Generate & Export (~30 mnt)", [
        ("GURU — RPP AI & Manajemen", guru_fill, [
            ('Guru', 'RPP AI', 'Buka fitur Generate RPP dengan AI'),
            ('Guru', 'RPP AI', 'Pilih kelas, mata pelajaran, topik'),
            ('Guru', 'RPP AI', 'Generate RPP AI — tunggu proses async'),
            ('Guru', 'RPP AI', 'Verifikasi loading/polling berjalan'),
            ('Guru', 'RPP AI', 'Verifikasi hasil RPP AI tampil lengkap'),
            ('Guru', 'RPP', 'Export RPP — verifikasi file terunduh/terbuka'),
            ('Guru', 'RPP', 'Hapus RPP, verifikasi hilang dari list'),
            ('Guru', 'RPP', 'Cek status approval RPP (approved/rejected/pending)'),
            ('Guru', 'RPP', 'Lihat RPP yang sudah di-approve vs yang pending'),
            ('Guru', 'RPP', 'Generate RPP AI kedua dengan topik berbeda'),
        ]),
    ]),
    ("HARI 9 — Input Nilai & Rekap Nilai (~30 mnt)", [
        ("GURU — Nilai Siswa", guru_fill, [
            ('Guru', 'Nilai', 'Buka halaman Input Nilai'),
            ('Guru', 'Nilai', 'Pilih kelas dan mata pelajaran'),
            ('Guru', 'Nilai', 'Input nilai untuk 5+ siswa (berbagai tipe penilaian)'),
            ('Guru', 'Nilai', 'Simpan nilai'),
            ('Guru', 'Nilai', 'Verifikasi nilai tersimpan dengan benar'),
            ('Guru', 'Nilai', 'Edit nilai salah satu siswa'),
            ('Guru', 'Nilai', 'Simpan ulang & verifikasi perubahan'),
            ('Guru', 'Rekap', 'Buka halaman Rekap Nilai'),
            ('Guru', 'Rekap', 'Verifikasi data rekap sesuai nilai yang diinput'),
            ('Guru', 'Rekap', 'Filter rekap berdasarkan mata pelajaran'),
            ('Guru', 'Rekap', 'Filter berdasarkan tahun ajaran'),
        ]),
    ]),
    ("HARI 10 — Raport & Rekomendasi AI (~30 mnt)", [
        ("GURU — Raport", guru_fill, [
            ('Guru', 'Raport', 'Buka halaman Raport'),
            ('Guru', 'Raport', 'Pilih kelas dan semester'),
            ('Guru', 'Raport', 'Lihat detail raport per siswa'),
            ('Guru', 'Raport', 'Cetak/print raport — verifikasi output'),
        ]),
        ("GURU WALI KELAS (T1 & T5) — Rekomendasi AI", guru_fill, [
            ('Guru', 'Rekomendasi', 'Buka halaman Rekomendasi Pembelajaran'),
            ('Guru', 'Rekomendasi', 'Generate rekomendasi AI untuk KELAS'),
            ('Guru', 'Rekomendasi', 'Tunggu async job, verifikasi hasil tampil'),
            ('Guru', 'Rekomendasi', 'Generate rekomendasi AI untuk SISWA individual'),
            ('Guru', 'Rekomendasi', 'Lihat detail rekomendasi'),
            ('Guru', 'Rekomendasi', 'Edit rekomendasi'),
            ('Guru', 'Rekomendasi', 'Ubah status: pending → in_progress → completed'),
            ('Guru', 'Rekomendasi', 'Dismiss rekomendasi'),
        ]),
    ]),
    ("HARI 11 — Wali: Aktivitas Kelas & Nilai Anak (~30 mnt)", [
        ("WALI — Monitoring Aktivitas", wali_fill, [
            ('Wali', 'Login', 'Login/switch ke akun Wali'),
            ('Wali', 'Aktivitas', 'Buka halaman Aktivitas Kelas'),
            ('Wali', 'Aktivitas', 'Verifikasi aktivitas yang dibuat guru tampil'),
            ('Wali', 'Aktivitas', 'Filter berdasarkan tanggal'),
            ('Wali', 'Aktivitas', 'Filter berdasarkan mata pelajaran'),
            ('Wali', 'Aktivitas', 'Verifikasi badge unread hilang setelah dibaca'),
            ('Wali', 'Aktivitas', 'Cek tugas/PR tampil dengan deadline yang benar'),
            ('Wali', 'Multi-Anak', 'T3 & T6: Switch anak, verifikasi data aktivitas berubah'),
        ]),
        ("WALI — Monitoring Nilai", wali_fill, [
            ('Wali', 'Nilai', 'Buka halaman Nilai'),
            ('Wali', 'Nilai', 'Verifikasi nilai yang diinput guru tampil'),
            ('Wali', 'Nilai', 'Filter berdasarkan mata pelajaran'),
            ('Wali', 'Nilai', 'Filter berdasarkan tahun ajaran'),
            ('Wali', 'Nilai', 'Verifikasi badge unread untuk nilai baru'),
            ('Wali', 'Multi-Anak', 'T3 & T6: Switch anak, verifikasi nilai berubah'),
        ]),
    ]),
    ("HARI 12 — Wali: Kehadiran & Billing (~30 mnt)", [
        ("WALI — Kehadiran Anak", wali_fill, [
            ('Wali', 'Kehadiran', 'Buka halaman Kehadiran'),
            ('Wali', 'Kehadiran', 'Verifikasi data absensi sesuai input guru'),
            ('Wali', 'Kehadiran', 'Cek rekap bulanan kehadiran'),
            ('Wali', 'Kehadiran', 'Filter berdasarkan rentang tanggal'),
            ('Wali', 'Kehadiran', 'Verifikasi statistik H/S/I/A benar'),
            ('Wali', 'Multi-Anak', 'T3 & T6: Switch anak, verifikasi kehadiran berubah'),
        ]),
        ("WALI — Billing/Tagihan", wali_fill, [
            ('Wali', 'Billing', 'Buka halaman Billing/Tagihan'),
            ('Wali', 'Billing', 'Verifikasi tagihan tampil dengan benar'),
            ('Wali', 'Billing', 'Cek status pembayaran (lunas/belum)'),
            ('Wali', 'Billing', 'Lihat riwayat pembayaran'),
            ('Wali', 'Billing', 'Verifikasi badge unread tagihan baru'),
            ('Wali', 'Multi-Anak', 'T3 & T6: Switch anak, verifikasi billing berubah'),
        ]),
    ]),
    ("HARI 13 — E-Raport Wali & Cross-Check Guru↔Wali (~30 mnt)", [
        ("WALI — E-Raport", wali_fill, [
            ('Wali', 'Raport', 'Buka halaman E-Raport'),
            ('Wali', 'Raport', 'Pilih semester/tahun ajaran'),
            ('Wali', 'Raport', 'Lihat raport lengkap anak'),
            ('Wali', 'Raport', 'Verifikasi data sesuai nilai dari guru'),
            ('Wali', 'Raport', 'Download/view detail raport'),
            ('Wali', 'Multi-Anak', 'T3 & T6: Switch anak, verifikasi raport berubah'),
        ]),
        ("CROSS-CHECK: Guru → Wali (Sekolah A & B)", crosscheck_fill, [
            ('Guru+Wali', 'Cross-check', 'Guru buat aktivitas baru → Wali cek tampil'),
            ('Guru+Wali', 'Cross-check', 'Guru input nilai baru → Wali cek nilai muncul'),
            ('Guru+Wali', 'Cross-check', 'Guru input absensi → Wali cek kehadiran update'),
            ('Guru+Wali', 'Cross-check', 'Guru buat data → Wali cek badge unread muncul'),
        ]),
        ("ISOLASI DATA ANTAR SEKOLAH", crosscheck_fill, [
            ('Guru+Wali', 'Isolasi', 'Data Sekolah A TIDAK tampil saat login Wali Sekolah B'),
            ('Guru+Wali', 'Isolasi', 'Data Sekolah B TIDAK tampil saat login Wali Sekolah A'),
            ('Wali', 'Multi-Anak', 'T3 & T6: Verifikasi data antar anak tidak tercampur'),
        ]),
    ]),
    ("HARI 14 — Edge Cases & Rekapitulasi (~30 mnt)", [
        ("EDGE CASES — Konektivitas & State", edge_fill, [
            ('Semua', 'Edge Case', 'Buka app tanpa koneksi internet — cek cache/offline'),
            ('Semua', 'Edge Case', 'Minimize app lalu buka lagi — state tetap'),
            ('Semua', 'Edge Case', 'Switch role Guru↔Wali berulang kali dengan cepat'),
            ('Semua', 'Edge Case', 'Buka 2 sekolah berbeda secara bergantian'),
        ]),
        ("EDGE CASES — Input & Batas", edge_fill, [
            ('Guru', 'Edge Case', 'Input nilai batas: 0, 100, 99.5'),
            ('Guru', 'Edge Case', 'Input nilai invalid: -1, 101, huruf'),
            ('Guru', 'Edge Case', 'Buat aktivitas dengan teks sangat panjang (500+ char)'),
            ('Guru', 'Edge Case', 'Input karakter spesial: !@#$%^&*(){}[]'),
            ('Guru', 'Edge Case', 'Scroll cepat pada list panjang — tidak crash'),
        ]),
        ("EDGE CASES — AI & Performa", edge_fill, [
            ('Guru', 'Edge Case', 'Generate AI saat koneksi lambat — verifikasi timeout'),
            ('Guru', 'Edge Case', 'Batalkan/keluar saat proses AI sedang berjalan'),
            ('Guru', 'Edge Case', 'Generate AI berulang kali — cek rate limiting'),
        ]),
        ("REKAPITULASI", recap_fill, [
            ('Semua', 'Rekap', 'Review semua bug yang ditemukan selama 14 hari'),
            ('Semua', 'Rekap', 'Verifikasi ulang bug Critical & Major'),
            ('Semua', 'Rekap', 'Isi form rekapitulasi di sheet Laporan Bug'),
            ('Semua', 'Rekap', 'Submit laporan akhir testing'),
        ]),
    ]),
]

row = 2
no = 1
for day_title, sections in days_gw:
    ws2.merge_cells(f'A{row}:{get_column_letter(MAX_GW)}{row}')
    ws2.cell(row=row, column=1, value=day_title)
    style_day_row(ws2, row, MAX_GW)
    ws2.row_dimensions[row].height = 28
    row += 1
    for section_title, fill, tasks in sections:
        ws2.merge_cells(f'A{row}:{get_column_letter(MAX_GW)}{row}')
        ws2.cell(row=row, column=1, value=section_title)
        style_section(ws2, row, MAX_GW)
        row += 1
        for role, kategori, task in tasks:
            ws2.cell(row=row, column=1, value=no)
            ws2.cell(row=row, column=2, value=role)
            ws2.cell(row=row, column=3, value=kategori)
            ws2.cell(row=row, column=4, value=task)
            for c in range(5, MAX_GW):
                ws2.cell(row=row, column=c, value='⬜')
            apply_row(ws2, row, MAX_GW, font=normal_font, fill=fill)
            ws2.cell(row=row, column=1).alignment = center
            ws2.cell(row=row, column=2).alignment = center
            for c in range(5, MAX_GW):
                ws2.cell(row=row, column=c).alignment = center
            no += 1
            row += 1
    row += 1

ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 58
for c_letter in ['E','F','G','H','I','J','K','L','M']:
    ws2.column_dimensions[c_letter].width = 6
ws2.column_dimensions['N'].width = 28
ws2.freeze_panes = 'A2'

# Add dropdown validation for checkbox columns (T1-T9) so testers can click to check
dv_gw = DataValidation(type="list", formula1='"✅,⬜"', allow_blank=True)
dv_gw.prompt = "Klik untuk pilih status"
dv_gw.promptTitle = "Status"
ws2.add_data_validation(dv_gw)
for c in range(5, MAX_GW):  # columns E to M (T1-T9)
    col_letter = get_column_letter(c)
    dv_gw.add(f'{col_letter}2:{col_letter}1048576')


# ============================================================
# SHEET 3: JADWAL 14 HARI — ADMIN (T10-T12)
# ============================================================
ws3 = wb.create_sheet("Jadwal Admin (T10-T12)")
ws3.sheet_properties.tabColor = "2E86C1"

cols_admin = ['No', 'Role', 'Kategori', 'Task Testing', 'T10', 'T11', 'T12', 'Catatan Bug']
MAX_AD = len(cols_admin)
for col, h in enumerate(cols_admin, 1):
    ws3.cell(row=1, column=col, value=h)
style_header_row(ws3, 1, MAX_AD)

days_admin = [
    ("HARI 1 — Login, Dashboard Admin & Pengumuman (~30 mnt)", [
        ("ADMIN — Dashboard & Overview", admin_fill, [
            ('Admin', 'Login', 'Login dengan akun Admin'),
            ('Admin', 'Dashboard', 'Verifikasi dashboard Admin tampil (cards: Data, Jadwal, Nilai, dll)'),
            ('Admin', 'Dashboard', 'Cek statistik: total siswa, guru, kelas'),
            ('Admin', 'Dashboard', 'Cek statistik keuangan di dashboard'),
            ('Admin', 'Dashboard', 'Cek badge unread pada setiap menu admin'),
            ('Admin', 'Pengumuman', 'Buka halaman Pengumuman admin'),
            ('Admin', 'Navigasi', 'Buka setiap menu admin, verifikasi navigasi lancar'),
            ('Admin', 'Settings', 'Ganti bahasa (ID ↔ EN), verifikasi teks berubah'),
        ]),
        ("T10 & T11: GURU + WALI Dashboard (di sekolah lainnya)", guru_fill, [
            ('Guru', 'Login', 'Switch ke akun Guru, verifikasi dashboard Guru'),
            ('Wali', 'Login', 'Switch ke akun Wali, verifikasi dashboard Wali'),
            ('Wali', 'Dashboard', 'Cek badge unread dan statistik Wali'),
        ]),
    ]),
    ("HARI 2 — Admin: Manajemen Siswa (~30 mnt)", [
        ("ADMIN — CRUD Siswa", admin_fill, [
            ('Admin', 'Siswa', 'Buka halaman Manajemen Siswa'),
            ('Admin', 'Siswa', 'Lihat daftar siswa dengan pagination'),
            ('Admin', 'Siswa', 'Search siswa berdasarkan nama'),
            ('Admin', 'Siswa', 'Filter siswa berdasarkan KELAS'),
            ('Admin', 'Siswa', 'Filter berdasarkan STATUS (aktif/nonaktif)'),
            ('Admin', 'Siswa', 'Filter berdasarkan JENIS KELAMIN'),
            ('Admin', 'Siswa', 'Tambah siswa baru (isi semua field)'),
            ('Admin', 'Siswa', 'Edit data siswa yang sudah ada'),
            ('Admin', 'Siswa', 'Lihat detail profil siswa'),
            ('Admin', 'Siswa', 'Assign wali/guardian ke siswa'),
            ('Admin', 'Import', 'Download template Excel import siswa'),
            ('Admin', 'Import', 'Import siswa dari file Excel (batch)'),
            ('Admin', 'Siswa', 'Hapus siswa (verifikasi konfirmasi muncul)'),
        ]),
    ]),
    ("HARI 3 — Admin: Manajemen Guru (~30 mnt)", [
        ("ADMIN — CRUD Guru", admin_fill, [
            ('Admin', 'Guru', 'Buka halaman Manajemen Guru'),
            ('Admin', 'Guru', 'Lihat daftar guru dengan pagination'),
            ('Admin', 'Guru', 'Search guru berdasarkan nama'),
            ('Admin', 'Guru', 'Tambah guru baru (isi semua field)'),
            ('Admin', 'Guru', 'Assign mata pelajaran ke guru'),
            ('Admin', 'Guru', 'Remove mata pelajaran dari guru'),
            ('Admin', 'Guru', 'Edit data guru yang sudah ada'),
            ('Admin', 'Guru', 'Lihat detail guru (kelas & mapel assigned)'),
            ('Admin', 'Import', 'Download template Excel import guru'),
            ('Admin', 'Import', 'Import guru dari file Excel (batch)'),
            ('Admin', 'Guru', 'Hapus guru (verifikasi konfirmasi muncul)'),
        ]),
    ]),
    ("HARI 4 — Admin: Manajemen Kelas (~30 mnt)", [
        ("ADMIN — CRUD Kelas", admin_fill, [
            ('Admin', 'Kelas', 'Buka halaman Manajemen Kelas'),
            ('Admin', 'Kelas', 'Lihat daftar kelas dengan pagination'),
            ('Admin', 'Kelas', 'Tambah kelas baru'),
            ('Admin', 'Kelas', 'Edit data kelas'),
            ('Admin', 'Kelas', 'Assign wali kelas (homeroom teacher)'),
            ('Admin', 'Kelas', 'Lihat siswa per kelas'),
            ('Admin', 'Import', 'Import kelas dari Excel'),
            ('Admin', 'Export', 'Export data kelas ke Excel'),
            ('Admin', 'Kelas', 'Hapus kelas (verifikasi konfirmasi)'),
            ('Admin', 'Promosi', 'Buka Wizard Kenaikan Kelas'),
            ('Admin', 'Promosi', 'Pilih kelas asal dan kelas tujuan'),
            ('Admin', 'Promosi', 'Pilih siswa untuk dipromosikan'),
            ('Admin', 'Promosi', 'Assign guru ke kelas baru'),
        ]),
    ]),
    ("HARI 5 — Admin: Mata Pelajaran & Jadwal (~30 mnt)", [
        ("ADMIN — Manajemen Mata Pelajaran", admin_fill, [
            ('Admin', 'Mapel', 'Buka halaman Manajemen Mata Pelajaran'),
            ('Admin', 'Mapel', 'Tambah mata pelajaran baru'),
            ('Admin', 'Mapel', 'Edit mata pelajaran'),
            ('Admin', 'Mapel', 'Attach kelas ke mata pelajaran'),
            ('Admin', 'Mapel', 'Detach kelas dari mata pelajaran'),
            ('Admin', 'Mapel', 'Import mapel dari Excel'),
            ('Admin', 'Mapel', 'Hapus mata pelajaran'),
        ]),
        ("ADMIN — Manajemen Jadwal", admin_fill, [
            ('Admin', 'Jadwal', 'Buka halaman Manajemen Jadwal'),
            ('Admin', 'Jadwal', 'Tambah jadwal mengajar baru'),
            ('Admin', 'Jadwal', 'Edit jadwal yang sudah ada'),
            ('Admin', 'Jadwal', 'Filter jadwal per guru, kelas, mapel'),
            ('Admin', 'Jadwal', 'Import jadwal dari Excel (cek conflict detection)'),
            ('Admin', 'Jadwal', 'Export jadwal ke Excel'),
            ('Admin', 'Jadwal', 'Hapus jadwal'),
            ('Admin', 'Jam Pel', 'Kelola Jam Pelajaran (tambah/hapus session)'),
        ]),
    ]),
    ("HARI 6 — Admin: Kelola Pengumuman (~30 mnt)", [
        ("ADMIN — CRUD Pengumuman", admin_fill, [
            ('Admin', 'Pengumuman', 'Buka halaman Manajemen Pengumuman'),
            ('Admin', 'Pengumuman', 'Buat pengumuman baru — prioritas TINGGI'),
            ('Admin', 'Pengumuman', 'Buat pengumuman — prioritas SEDANG, target: Guru'),
            ('Admin', 'Pengumuman', 'Buat pengumuman — prioritas RENDAH, target: Wali'),
            ('Admin', 'Pengumuman', 'Buat pengumuman dengan LAMPIRAN FILE'),
            ('Admin', 'Pengumuman', 'Edit pengumuman yang sudah ada'),
            ('Admin', 'Pengumuman', 'Filter pengumuman per prioritas & target role'),
            ('Admin', 'Pengumuman', 'Hapus pengumuman'),
            ('Admin', 'Pengumuman', 'Verifikasi pengumuman tampil di sisi Guru & Wali'),
        ]),
    ]),
    ("HARI 7 — Admin: Laporan Absensi (~30 mnt)", [
        ("ADMIN — Report Absensi", admin_fill, [
            ('Admin', 'Absensi', 'Buka halaman Laporan Absensi'),
            ('Admin', 'Absensi', 'Lihat data absensi dengan pagination'),
            ('Admin', 'Absensi', 'Filter absensi berdasarkan GURU'),
            ('Admin', 'Absensi', 'Filter berdasarkan KELAS'),
            ('Admin', 'Absensi', 'Filter berdasarkan STATUS'),
            ('Admin', 'Absensi', 'Filter berdasarkan TANGGAL'),
            ('Admin', 'Absensi', 'Lihat detail absensi per siswa'),
            ('Admin', 'Absensi', 'Lihat statistik & chart absensi'),
            ('Admin', 'Export', 'Export absensi ke Excel (1 bulan)'),
            ('Admin', 'Export', 'Export absensi multi-bulan'),
            ('Admin', 'Absensi', 'Hapus record absensi (verifikasi konfirmasi)'),
            ('Admin', 'Aktivitas', 'Buka Aktivitas Kelas admin — lihat & filter aktivitas'),
        ]),
    ]),
    ("HARI 8 — Admin: RPP Approval & Nilai Admin (~30 mnt)", [
        ("ADMIN — RPP Approval", admin_fill, [
            ('Admin', 'RPP', 'Buka halaman Manajemen RPP'),
            ('Admin', 'RPP', 'Lihat semua RPP dari guru dengan pagination'),
            ('Admin', 'RPP', 'Filter RPP berdasarkan GURU'),
            ('Admin', 'RPP', 'Filter berdasarkan STATUS (Pending/Approved/Rejected)'),
            ('Admin', 'RPP', 'Search RPP'),
            ('Admin', 'RPP', 'Lihat detail RPP untuk review'),
            ('Admin', 'RPP', 'Approve RPP yang Pending'),
            ('Admin', 'RPP', 'Reject RPP yang Pending (beri alasan)'),
            ('Admin', 'RPP', 'Export data RPP'),
        ]),
        ("ADMIN — Nilai", admin_fill, [
            ('Admin', 'Nilai', 'Buka halaman Input Nilai (admin)'),
            ('Admin', 'Nilai', 'Lihat nilai per mapel & kelas'),
            ('Admin', 'Nilai', 'Export nilai ke Excel'),
            ('Admin', 'Nilai', 'Verifikasi data nilai sesuai input guru'),
        ]),
    ]),
    ("HARI 9 — Admin: Raport & Keuangan (Jenis Pembayaran) (~30 mnt)", [
        ("ADMIN — Raport Management", admin_fill, [
            ('Admin', 'Raport', 'Buka halaman Manajemen Raport'),
            ('Admin', 'Raport', 'Filter raport per kelas & tahun ajaran'),
            ('Admin', 'Raport', 'Lihat detail raport siswa'),
            ('Admin', 'Raport', 'Publish raport — buat tersedia untuk wali'),
            ('Admin', 'Raport', 'Download raport PDF per siswa'),
            ('Admin', 'Raport', 'Export raport batch ke Excel'),
        ]),
        ("ADMIN — Keuangan: Jenis Pembayaran", admin_fill, [
            ('Admin', 'Keuangan', 'Buka halaman Keuangan/Finance'),
            ('Admin', 'Keuangan', 'Lihat dashboard statistik keuangan'),
            ('Admin', 'Keuangan', 'Buat jenis pembayaran baru (SPP, Buku, dll)'),
            ('Admin', 'Keuangan', 'Set periode pembayaran (bulanan/tahunan)'),
            ('Admin', 'Keuangan', 'Set status jenis pembayaran (aktif/nonaktif)'),
            ('Admin', 'Keuangan', 'Hapus jenis pembayaran'),
        ]),
    ]),
    ("HARI 10 — Admin: Keuangan (Tagihan) (~30 mnt)", [
        ("ADMIN — Generate & Kelola Tagihan", admin_fill, [
            ('Admin', 'Tagihan', 'Buka halaman Tagihan/Bills'),
            ('Admin', 'Tagihan', 'Lihat semua tagihan siswa dengan pagination'),
            ('Admin', 'Tagihan', 'Filter tagihan per KELAS'),
            ('Admin', 'Tagihan', 'Filter per SISWA'),
            ('Admin', 'Tagihan', 'Filter per STATUS (lunas/belum/sebagian)'),
            ('Admin', 'Tagihan', 'Filter per PERIODE'),
            ('Admin', 'Tagihan', 'Search tagihan'),
            ('Admin', 'Tagihan', 'Generate tagihan untuk seluruh kelas'),
            ('Admin', 'Tagihan', 'Verifikasi tagihan ter-generate dengan benar'),
            ('Admin', 'Tagihan', 'Export data tagihan'),
            ('Admin', 'Tagihan', 'Hapus tagihan yang sudah di-generate'),
            ('Admin', 'Tagihan', 'Lihat laporan tagihan bulanan'),
        ]),
    ]),
    ("HARI 11 — Admin: Keuangan (Pembayaran) + Wali T10-T11 (~30 mnt)", [
        ("ADMIN — Pembayaran & Verifikasi", admin_fill, [
            ('Admin', 'Bayar', 'Buka halaman Pembayaran'),
            ('Admin', 'Bayar', 'Lihat pembayaran pending verifikasi'),
            ('Admin', 'Bayar', 'Filter status pembayaran'),
            ('Admin', 'Bayar', 'Input pembayaran MANUAL untuk siswa'),
            ('Admin', 'Bayar', 'Verifikasi pembayaran (approve)'),
            ('Admin', 'Bayar', 'Lihat riwayat pembayaran'),
            ('Admin', 'Bayar', 'Cek statistik: total revenue, belum bayar, sebagian'),
            ('Admin', 'Laporan', 'Lihat laporan keuangan per kelas'),
        ]),
        ("T10 & T11: WALI — Monitoring Anak (di sekolah lainnya)", wali_fill, [
            ('Wali', 'Aktivitas', 'Buka Aktivitas Kelas, verifikasi data tampil'),
            ('Wali', 'Nilai', 'Buka Nilai, verifikasi tampil'),
            ('Wali', 'Kehadiran', 'Buka Kehadiran, verifikasi data'),
            ('Wali', 'Billing', 'Buka Billing, cek tagihan'),
        ]),
    ]),
    ("HARI 12 — Admin: Settings Sekolah + Wali T10-T11 (~30 mnt)", [
        ("ADMIN — Settings Sekolah", admin_fill, [
            ('Admin', 'Settings', 'Buka halaman Settings Sekolah'),
            ('Admin', 'Settings', 'Edit nama sekolah'),
            ('Admin', 'Settings', 'Edit alamat sekolah'),
            ('Admin', 'Settings', 'Set jenjang sekolah (SD/SMP/SMA/SMK)'),
            ('Admin', 'Waktu', 'Buka Settings Waktu'),
            ('Admin', 'Waktu', 'Kelola hari sekolah'),
            ('Admin', 'Waktu', 'Tambah sesi jam pelajaran'),
            ('Admin', 'Waktu', 'Hapus sesi jam pelajaran'),
            ('Admin', 'Semester', 'Kelola semester & tahun ajaran'),
            ('Admin', 'Profil', 'Update profil admin (nama, telp, alamat)'),
            ('Admin', 'Profil', 'Ganti password admin'),
        ]),
        ("T10 & T11: WALI — E-Raport Anak", wali_fill, [
            ('Wali', 'Raport', 'Buka E-Raport, lihat raport anak'),
            ('Wali', 'Raport', 'Download/view detail raport'),
        ]),
    ]),
    ("HARI 13 — Admin: Cross-Check Semua Role (~30 mnt)", [
        ("ADMIN — Cross-Check dengan Guru & Wali", crosscheck_fill, [
            ('Admin', 'Cross-check', 'Admin buat pengumuman → Guru & Wali cek tampil'),
            ('Admin', 'Cross-check', 'Admin publish raport → Wali cek E-Raport muncul'),
            ('Admin', 'Cross-check', 'Admin generate tagihan → Wali cek Billing muncul'),
            ('Admin', 'Cross-check', 'Admin approve RPP → Guru cek status berubah'),
            ('Admin', 'Cross-check', 'Admin reject RPP → Guru cek status berubah'),
            ('Admin', 'Cross-check', 'Admin tambah siswa → Guru cek siswa muncul di kelas'),
            ('Admin', 'Cross-check', 'Admin ubah jadwal → Guru cek jadwal berubah'),
            ('Admin', 'Cross-check', 'Admin input pembayaran → Wali cek status billing update'),
        ]),
        ("ISOLASI DATA — Admin Antar Sekolah (T12)", crosscheck_fill, [
            ('Admin', 'Isolasi', 'Data admin Sekolah A TIDAK tampil di Sekolah B'),
            ('Admin', 'Isolasi', 'Data admin Sekolah B TIDAK tampil di Sekolah A'),
            ('Admin', 'Isolasi', 'Siswa Sekolah A tidak muncul di kelas Sekolah B'),
            ('Admin', 'Isolasi', 'Guru Sekolah A tidak muncul di jadwal Sekolah B'),
        ]),
        ("T10 & T11: GURU + WALI Cross-Check", crosscheck_fill, [
            ('Guru+Wali', 'Cross-check', 'Guru buat aktivitas → switch Wali, cek tampil'),
            ('Guru+Wali', 'Cross-check', 'Guru input nilai → switch Wali, cek nilai muncul'),
            ('Guru+Wali', 'Cross-check', 'Guru input absensi → switch Wali, cek kehadiran'),
        ]),
    ]),
    ("HARI 14 — Edge Cases & Rekapitulasi (~30 mnt)", [
        ("EDGE CASES — Admin", edge_fill, [
            ('Admin', 'Edge Case', 'Import Excel dengan data invalid/duplikat'),
            ('Admin', 'Edge Case', 'Import jadwal dengan conflict (overlap waktu)'),
            ('Admin', 'Edge Case', 'Generate tagihan untuk kelas tanpa siswa'),
            ('Admin', 'Edge Case', 'Hapus guru yang masih punya jadwal aktif'),
            ('Admin', 'Edge Case', 'Hapus kelas yang masih punya siswa'),
            ('Admin', 'Edge Case', 'Input pembayaran melebihi total tagihan'),
            ('Admin', 'Edge Case', 'Edit settings sekolah dengan data kosong/invalid'),
            ('Admin', 'Edge Case', 'Promosi kelas ke tahun ajaran yang sama'),
        ]),
        ("EDGE CASES — Umum (Semua T10-T12)", edge_fill, [
            ('Semua', 'Edge Case', 'App tanpa internet — cek cache/offline'),
            ('Semua', 'Edge Case', 'Switch role Admin→Guru→Wali berulang kali'),
            ('Semua', 'Edge Case', 'Minimize app lalu buka lagi — state tetap'),
        ]),
        ("REKAPITULASI", recap_fill, [
            ('Semua', 'Rekap', 'Review semua bug yang ditemukan selama 14 hari'),
            ('Semua', 'Rekap', 'Verifikasi ulang bug Critical & Major'),
            ('Semua', 'Rekap', 'Isi form rekapitulasi di sheet Laporan Bug'),
            ('Semua', 'Rekap', 'Submit laporan akhir testing'),
        ]),
    ]),
]

row = 2
no = 1
for day_title, sections in days_admin:
    ws3.merge_cells(f'A{row}:{get_column_letter(MAX_AD)}{row}')
    ws3.cell(row=row, column=1, value=day_title)
    style_day_row(ws3, row, MAX_AD)
    ws3.row_dimensions[row].height = 28
    row += 1
    for section_title, fill, tasks in sections:
        ws3.merge_cells(f'A{row}:{get_column_letter(MAX_AD)}{row}')
        ws3.cell(row=row, column=1, value=section_title)
        style_section(ws3, row, MAX_AD)
        row += 1
        for role, kategori, task in tasks:
            ws3.cell(row=row, column=1, value=no)
            ws3.cell(row=row, column=2, value=role)
            ws3.cell(row=row, column=3, value=kategori)
            ws3.cell(row=row, column=4, value=task)
            for c in range(5, MAX_AD):
                ws3.cell(row=row, column=c, value='⬜')
            apply_row(ws3, row, MAX_AD, font=normal_font, fill=fill)
            ws3.cell(row=row, column=1).alignment = center
            ws3.cell(row=row, column=2).alignment = center
            for c in range(5, MAX_AD):
                ws3.cell(row=row, column=c).alignment = center
            no += 1
            row += 1
    row += 1

ws3.column_dimensions['A'].width = 5
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 58
ws3.column_dimensions['E'].width = 8
ws3.column_dimensions['F'].width = 8
ws3.column_dimensions['G'].width = 8
ws3.column_dimensions['H'].width = 28
ws3.freeze_panes = 'A2'

# Add dropdown validation for checkbox columns (T10-T12) so testers can click to check
dv_ad = DataValidation(type="list", formula1='"✅,⬜"', allow_blank=True)
dv_ad.prompt = "Klik untuk pilih status"
dv_ad.promptTitle = "Status"
ws3.add_data_validation(dv_ad)
for c in range(5, MAX_AD):  # columns E to G (T10-T12)
    col_letter = get_column_letter(c)
    dv_ad.add(f'{col_letter}2:{col_letter}1048576')


# ============================================================
# SHEET 4: RINGKASAN PER TESTER
# ============================================================
ws4 = wb.create_sheet("Ringkasan Per Tester")
ws4.sheet_properties.tabColor = "27AE60"

rt_cols = ['Hari', 'Tema', 'T1-T9 (Guru+Wali)', 'T10 (Admin+Guru+Wali)', 'T11 (Admin+Guru+Wali)', 'T12 (Admin)', 'Estimasi']
for col, h in enumerate(rt_cols, 1):
    ws4.cell(row=1, column=col, value=h)
style_header_row(ws4, 1, 7)

summary = [
    ['Hari 1', 'Login & Dashboard', 'Guru+Wali Dashboard', 'Admin+Guru+Wali Dashboard', 'Admin+Guru+Wali Dashboard', 'Admin Dashboard A+B', '~30 mnt'],
    ['Hari 2', 'Jadwal / Siswa', 'Guru: Jadwal Mengajar', 'Admin: CRUD Siswa (A)', 'Admin: CRUD Siswa (B)', 'Admin: CRUD Siswa (A+B)', '~30 mnt'],
    ['Hari 3', 'Aktivitas / Guru', 'Guru: Buat Aktivitas', 'Admin: CRUD Guru (A)', 'Admin: CRUD Guru (B)', 'Admin: CRUD Guru (A+B)', '~30 mnt'],
    ['Hari 4', 'Kelola Aktivitas / Kelas', 'Guru: Edit/Hapus/Filter', 'Admin: CRUD Kelas (A)', 'Admin: CRUD Kelas (B)', 'Admin: CRUD Kelas (A+B)', '~30 mnt'],
    ['Hari 5', 'Absensi / Mapel+Jadwal', 'Guru: Input Absensi', 'Admin: Mapel+Jadwal (A)', 'Admin: Mapel+Jadwal (B)', 'Admin: Mapel+Jadwal (A+B)', '~30 mnt'],
    ['Hari 6', 'Materi / Pengumuman', 'Guru: CRUD Materi', 'Admin: Pengumuman (A)', 'Admin: Pengumuman (B)', 'Admin: Pengumuman (A+B)', '~30 mnt'],
    ['Hari 7', 'AI Materi+RPP / Absensi', 'Guru: AI Materi + RPP', 'Admin: Lap. Absensi (A)', 'Admin: Lap. Absensi (B)', 'Admin: Lap. Absensi (A+B)', '~30 mnt'],
    ['Hari 8', 'RPP AI / RPP Approval', 'Guru: RPP AI + Export', 'Admin: RPP Approval (A)', 'Admin: RPP Approval (B)', 'Admin: RPP Approval (A+B)', '~30 mnt'],
    ['Hari 9', 'Nilai / Raport+Keuangan', 'Guru: Input Nilai', 'Admin: Raport+Keuangan (A)', 'Admin: Raport+Keuangan (B)', 'Admin: Keuangan Jenis (A+B)', '~30 mnt'],
    ['Hari 10', 'Raport+Rekom / Tagihan', 'Guru: Raport + AI Rekom', 'Admin: Tagihan (A)', 'Admin: Tagihan (B)', 'Admin: Tagihan (A+B)', '~30 mnt'],
    ['Hari 11', 'Wali Aktivitas+Nilai / Bayar', 'Wali: Aktivitas + Nilai', 'Admin: Bayar (A) + Wali', 'Admin: Bayar (B) + Wali', 'Admin: Pembayaran (A+B)', '~30 mnt'],
    ['Hari 12', 'Wali Kehadiran+Billing / Settings', 'Wali: Kehadiran + Billing', 'Admin: Settings (A) + Wali', 'Admin: Settings (B) + Wali', 'Admin: Settings (A+B)', '~30 mnt'],
    ['Hari 13', 'Cross-check Semua', 'Wali: E-Raport + Cross-check', 'Admin+Guru+Wali Cross-check', 'Admin+Guru+Wali Cross-check', 'Admin Cross-check (A+B)', '~30 mnt'],
    ['Hari 14', 'Edge Cases & Rekap', 'Edge Cases + Rekap', 'Edge Cases + Rekap', 'Edge Cases + Rekap', 'Admin Edge Cases + Rekap', '~30 mnt'],
]

for i, s in enumerate(summary):
    row = 2 + i
    for col, val in enumerate(s, 1):
        ws4.cell(row=row, column=col, value=val)
    apply_row(ws4, row, 7, font=normal_font)
    ws4.cell(row=row, column=1).font = bold_font
    ws4.cell(row=row, column=1).alignment = center

row = 18
ws4.merge_cells(f'A{row}:G{row}')
ws4.cell(row=row, column=1, value="CATATAN PENTING").font = Font(name='Calibri', bold=True, size=12, color='1B4F72')

notes = [
    'T1-T4: Guru di Sekolah A, Wali di Sekolah B',
    'T5-T7: Wali di Sekolah A, Guru di Sekolah B',
    'T8: Guru di Sekolah A (Seni Budaya), Wali di Sekolah B',
    'T9: Wali di Sekolah A, Guru di Sekolah B (Penjaskes)',
    'T10: Admin+Guru di Sekolah A (B. Arab), Wali di Sekolah B',
    'T11: Wali di Sekolah A, Admin+Guru di Sekolah B (TIK)',
    'T12: Admin di KEDUA sekolah (A+B) — no Guru/Wali role',
    'T3 & T6 memiliki 2 anak — wajib test switching anak',
    'T1 & T5 = Wali Kelas (homeroom) — wajib test AI Rekomendasi',
    'Setiap menemukan bug, langsung catat di sheet Laporan Bug',
    'T12 fokus isolasi data admin antar 2 sekolah',
]

for i, n in enumerate(notes):
    row += 1
    ws4.cell(row=row, column=1, value=f'{i+1}.').font = bold_font
    ws4.cell(row=row, column=1).alignment = center
    ws4.merge_cells(f'B{row}:G{row}')
    ws4.cell(row=row, column=2, value=n).font = normal_font

ws4.column_dimensions['A'].width = 10
ws4.column_dimensions['B'].width = 28
ws4.column_dimensions['C'].width = 30
ws4.column_dimensions['D'].width = 32
ws4.column_dimensions['E'].width = 32
ws4.column_dimensions['F'].width = 30
ws4.column_dimensions['G'].width = 10


# ============================================================
# SHEET 5: LAPORAN BUG
# ============================================================
ws5 = wb.create_sheet("Laporan Bug")
ws5.sheet_properties.tabColor = "E74C3C"

bug_cols = ['No', 'Hari Ke-', 'Tanggal', 'Tester', 'Sekolah', 'Role', 'Halaman', 'Langkah Reproduksi', 'Expected', 'Actual', 'Severity', 'Screenshot', 'Status Fix']
for col, h in enumerate(bug_cols, 1):
    ws5.cell(row=1, column=col, value=h)
style_header_row(ws5, 1, len(bug_cols))

for r in range(2, 72):
    ws5.cell(row=r, column=1, value=r-1)
    for col in range(1, len(bug_cols)+1):
        ws5.cell(row=r, column=col).border = thin_border
        ws5.cell(row=r, column=col).font = normal_font
    ws5.cell(row=r, column=1).alignment = center

row = 73
ws5.cell(row=row, column=1, value="Panduan Severity:").font = bold_font
sev = [
    ('Critical', 'C0392B', 'App crash, data hilang, tidak bisa login, data bocor antar sekolah'),
    ('Major', 'E67E22', 'Fitur utama tidak berfungsi, data salah, badge tidak update'),
    ('Minor', 'F1C40F', 'Fitur berfungsi tapi tidak sempurna, UI minor glitch'),
    ('Cosmetic', '3498DB', 'Tampilan/UI tidak sesuai, typo, spacing salah'),
]
for label, color, desc in sev:
    row += 1
    ws5.cell(row=row, column=1, value=label).font = Font(name='Calibri', color=color, bold=True)
    ws5.merge_cells(f'B{row}:F{row}')
    ws5.cell(row=row, column=2, value=desc).font = normal_font

ws5.column_dimensions['A'].width = 5
ws5.column_dimensions['B'].width = 9
ws5.column_dimensions['C'].width = 12
ws5.column_dimensions['D'].width = 10
ws5.column_dimensions['E'].width = 10
ws5.column_dimensions['F'].width = 10
ws5.column_dimensions['G'].width = 16
ws5.column_dimensions['H'].width = 45
ws5.column_dimensions['I'].width = 28
ws5.column_dimensions['J'].width = 28
ws5.column_dimensions['K'].width = 10
ws5.column_dimensions['L'].width = 14
ws5.column_dimensions['M'].width = 12
ws5.freeze_panes = 'A2'

# ============================================================
# SAVE
# ============================================================
output = '/Users/macbook/development/projects/non-FCM/manajemennonfcm/Skenario_Testing_14Hari.xlsx'
wb.save(output)
print(f"Excel saved: {output}")
print(f"Sheets: {wb.sheetnames}")
